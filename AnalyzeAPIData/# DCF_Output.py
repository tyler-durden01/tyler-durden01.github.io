import openpyxl
from openpyxl import load_workbook
import os
import requests
from yahoo_fin.stock_info import *
import pandas as pd

# ticker = input("Enter ticker symbol: ")
ticker = 'ASML'

def duplicate_excel_file(original_file, duplicate_file):
    # Open the original Excel file
    wb = openpyxl.load_workbook(original_file)

    # Save a copy of the original Excel file with a new name
    wb.save(duplicate_file)

    print(f"Excel file duplicated successfully as '{duplicate_file}'.")

# Example usage
original_file = "DCF_Template.xlsx"
duplicate_file = f'{ticker}_DCF.xlsx'

duplicate_excel_file(original_file, duplicate_file)




api_key = '6fe8c4680cf2609b34c3674e0a32720b'

base_url = 'https://financialmodelingprep.com/api/v3/'

# Define the endpoints for both income statement and cash flow statement
income_statement_endpoint = f'income-statement/{ticker}'
balance_sheet_endpoint = f'balance-sheet-statement/{ticker}'
cash_flow_statement_endpoint = f'cash-flow-statement/{ticker}'
earnings_endpoint = f'historical/earning_calendar/{ticker}'

# Define the parameters for the requests
params = {
    'apikey': api_key,
    'period': 'quarter'
}

# Send a GET request to fetch the finanical statement data
income_response = requests.get(base_url + income_statement_endpoint, params=params)
cash_flow_response = requests.get(base_url + cash_flow_statement_endpoint, params=params)
balance_sheet_response = requests.get(base_url + balance_sheet_endpoint, params=params)
earnings_response = requests.get(base_url + earnings_endpoint, params=params)

# Check if both requests were successful (status code 200); can change this to a try and except; not a big fan of this loop
if income_response.status_code == 200 and cash_flow_response.status_code == 200:
    # Parse the JSON responses for both statements
    income_data = income_response.json()
    cash_flow_data = cash_flow_response.json()
    balance_sheet_data = balance_sheet_response.json()
    earnings_data = earnings_response.json()
    
    # Create Pandas DataFrames from the data
    income_df = pd.DataFrame(income_data)
    cash_flow_df = pd.DataFrame(cash_flow_data)
    balance_sheet_df = pd.DataFrame(balance_sheet_data)
    earnings_df = pd.DataFrame(earnings_data)

    # Filter and select the relevant columns for income statement
    income_df = income_df[[
        'date',
        'revenue',
        'grossProfit',
        # 'generalAndAdministrativeExpenses',
        # 'sellingAndMarketingExpenses',
        'sellingGeneralAndAdministrativeExpenses',
        'researchAndDevelopmentExpenses',
        'otherExpenses',
        'operatingIncome',
        'totalOtherIncomeExpensesNet',
        'interestExpense',
        'interestIncome',
        'incomeTaxExpense',
        'netIncome',
        ]]

    # Filter and select the relevant columns for cash flow statement
    cash_flow_df = cash_flow_df[[
        'date', 
        'freeCashFlow',
        'accountsReceivables',
        'accountsPayables',
        'inventory',
        'depreciationAndAmortization',
        'stockBasedCompensation',
        'deferredIncomeTax',
        'changeInWorkingCapital',
        'otherWorkingCapital',
        'otherNonCashItems',
        'netChangeInCash',
        'netCashProvidedByOperatingActivities',
        'capitalExpenditure',
        'investmentsInPropertyPlantAndEquipment',
        'acquisitionsNet',
        'purchasesOfInvestments',
        'netCashUsedForInvestingActivites',
        'debtRepayment',
        'commonStockIssued',
        'commonStockRepurchased',
        'dividendsPaid',
        'netCashUsedProvidedByFinancingActivities',
        'effectOfForexChangesOnCash',
        ]]
    
    balance_sheet_df = balance_sheet_df[[
        'date',
        'cashAndCashEquivalents',
        'shortTermInvestments',
        'cashAndShortTermInvestments',
        'netReceivables',
        'inventory',
        'otherCurrentAssets',
        'totalCurrentAssets',
        'propertyPlantEquipmentNet',
        'goodwill',
        'intangibleAssets',
        'longTermInvestments',
        'otherNonCurrentAssets',
        'totalNonCurrentAssets',
        'totalAssets',
        'accountPayables',
		'shortTermDebt',
		'taxPayables',
		'deferredRevenue',
		'otherCurrentLiabilities',
		'totalCurrentLiabilities',
		'longTermDebt',
		'deferredRevenueNonCurrent',
		'deferredTaxLiabilitiesNonCurrent',
		'otherNonCurrentLiabilities',
		'totalNonCurrentLiabilities',
		'otherLiabilities',
		'capitalLeaseObligations',
		'totalLiabilities',
		'preferredStock',
		'commonStock',
		'retainedEarnings',
		'accumulatedOtherComprehensiveIncomeLoss',
		'othertotalStockholdersEquity',
		'totalStockholdersEquity',
		'totalEquity',
		'totalLiabilitiesAndStockholdersEquity',
		'minorityInterest',
		'totalLiabilitiesAndTotalEquity',
		'totalInvestments',
		'totalDebt',
		'netDebt',
    ]]

    earnings_df = earnings_df.drop(columns=['symbol'])

else:
    print("Failed to retrieve data.")


# Merge the two DataFrames on the 'date' column to align the data
merged_df = pd.merge(income_df, cash_flow_df, on='date', how='inner')
merged_df = pd.merge(merged_df, balance_sheet_df, on='date', how='inner')

# print(merged_df)



# Convert numerical columns to numeric data types
numeric_columns = merged_df.columns.drop('date')
merged_df[numeric_columns] = merged_df[numeric_columns].apply(pd.to_numeric, errors='coerce')
merged_df.set_index('date', inplace=True)

merged_df = merged_df / 1000000
merged_df = merged_df.round(0)



merged_df = merged_df.transpose()
merged_df.rename(index={
    "revenue": "Revenue",
    "researchAndDevelopmentExpenses": "Research & Development",
    "grossProfit": "Gross Profit",
    "sellingGeneralAndAdministrativeExpenses": "Selling, General, & Administrative",
    "otherExpenses": "Other Operating Expenses",
    "incomeTaxExpense": "Taxes",
    "stockBasedCompensation": "Stock Based Compensation",
    "interestIncome": "Interest Income",
    "depreciationAndAmortization": "Depreciation & Amortization",
    "capitalExpenditure": "Capital Expenditure",
    "changeInWorkingCapital": "Change In Working Capital",
    }, inplace=True)



# wb = load_workbook(duplicate_file)

excel_writer = pd.ExcelWriter(duplicate_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
# excel_writer.book = duplicate_file


merged_df.to_excel(excel_writer=excel_writer, sheet_name="Python")


# For the Laptop
# excel_writer.save()

# For the Desktop
excel_writer.close()

# .applymap('{:,.0f}'.format)
# Convert 'date' column to datetime
# merged_df['date'] = pd.to_datetime(merged_df['date'])
# Print the final merged DataFrame
# print(merged_df)


'''
horizontal_df = merged_df.div(merged_df['revenue'], axis=0)
horizontal_df = horizontal_df.round(3)
print(horizontal_df)

# Calculate the % changes from the next period (descending order)
# quarterly_pct_change_df = merged_df[::-1].set_index('date').pct_change()
quarterly_pct_change_df = merged_df[::-1].pct_change()
quarterly_pct_change_df = (quarterly_pct_change_df[::-1]).round(3) # Reverse the result back to ascending order
print(quarterly_pct_change_df)

# yearly_pct_change_df = merged_df[::-1].set_index('date').pct_change(periods=4)
yearly_pct_change_df = merged_df[::-1].pct_change(periods=4)
yearly_pct_change_df = yearly_pct_change_df[::-1].round(3) # Reverse the result back to ascending order
print(yearly_pct_change_df)  

annual_df = merged_df.copy()
annual_df['year'] = pd.to_datetime(annual_df.index)
annual_df['year'] = annual_df['year'].dt.year
annual_df = annual_df.groupby('year').sum()
annual_df = annual_df[::-1]
annual_df.index.name = 'date'
print(annual_df)

annual_pct_change_df = annual_df[::-1].pct_change()
annual_pct_change_df = annual_pct_change_df[::-1].round(3)
print(annual_pct_change_df)

# Create a new DataFrame for cumulative sums
ttm_df = merged_df.copy()
# Calculate cumulative sums for each row in reverse order
for i in range(len(merged_df) - 4, -1, -1):
    ttm_df.iloc[i] = merged_df.iloc[i:i + 4].sum()
# Set NaN values for the last three rows
ttm_df.iloc[-3:] = None
# Print the cumulative sum DataFrame
print(ttm_df)

# yearly_pct_change_df = merged_df[::-1].set_index('date').pct_change(periods=4)
ttm_df_quarterly_pct_change = ttm_df[::-1].pct_change()
ttm_df_quarterly_pct_change = ttm_df_quarterly_pct_change[::-1].round(3) # Reverse the result back to ascending order
print(ttm_df_quarterly_pct_change)

# yearly_pct_change_df = merged_df[::-1].set_index('date').pct_change(periods=4)
ttm_df_yearly_pct_change = ttm_df[::-1].pct_change(periods=4)
ttm_df_yearly_pct_change = ttm_df_yearly_pct_change[::-1].round(3) # Reverse the result back to ascending order
print(ttm_df_yearly_pct_change)

# earnings_df.set_index('date', inplace=True)
print(earnings_df)
# need to format this df so that date is the index




# Create a Pandas Excel writer using XlsxWriter as the engine.
excel_file_name = f'{ticker}_financial_data.xlsx'
excel_file_path = os.path.join(folder_path, excel_file_name)
excel_writer = pd.ExcelWriter(excel_file_path, engine='xlsxwriter')


# Write each DataFrame to a different sheet in the Excel file.
merged_df.to_excel(excel_writer, sheet_name='Merged Data', index=True)
horizontal_df.to_excel(excel_writer, sheet_name='Margins', index=True)
quarterly_pct_change_df.to_excel(excel_writer, sheet_name='Quarterly % Change', index=True)
yearly_pct_change_df.to_excel(excel_writer, sheet_name='Yearly % Change', index=True)
annual_df.to_excel(excel_writer, sheet_name='Annual Data', index=True)
annual_pct_change_df.to_excel(excel_writer, sheet_name='Annual % Change', index=True)
ttm_df.to_excel(excel_writer, sheet_name='TTM', index=True)
ttm_df_quarterly_pct_change.to_excel(excel_writer, sheet_name='TTM Quarterly % Change')
ttm_df_yearly_pct_change.to_excel(excel_writer, sheet_name='TTM Yearly % Change')
earnings_df.to_excel(excel_writer, sheet_name='Historical Earnings', index=False)
'''


























os.system(f'start excel "{duplicate_file}"')